import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import io

# --- Configuration ---
DATA_SHEET_TEMPLATE_PATH = "Data Sheet.xlsm"
DATA_SHEET_SHEET_NAME = "FORM"
DATA_SHEET_FORM_START_ROW = 9
MAX_ROWS_TO_CLEAR_IN_TEMPLATE = 200 
CALC_SHEET_SHEET_NAME = "PSV"

# --- Helper Functions ---
def convert_value(value):
    if pd.isna(value):
        return "-"
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value

def get_state(phase_value):
    if isinstance(phase_value, str):
        phase_value = phase_value.strip().upper()
        if phase_value == "V":
            return "VAPOR"
        elif phase_value == "S":
            return "STEAM" 
        elif phase_value == "L":
            return "LIQUID"
    return "-"

def format_back_pressure_calculated(row_data):
    min_bp = pd.to_numeric(row_data.get('Min. BP@Header'), errors='coerce')
    max_bp = pd.to_numeric(row_data.get('Max. BP@Header'), errors='coerce')

    if pd.isna(min_bp) or pd.isna(max_bp):
        return "-"
    
    try:
        if max_bp - min_bp >= 0:
            return f"{convert_value(min_bp)} / {convert_value(round(max_bp - min_bp, 1))}"
        else:
            return "-" 
    except Exception:
        return "-"

# --- Column Name Mappings ---
CALC_SHEET_HEADER_MAPPING = {
    "Tag No.": "Tag No.",
    "State (V/S/L)": "Phase", 
    "Fluid": "Fluid",
    "Flowing Fluid at Relieving Conditions": "Fluid", 
    "Flow Rate": "Flow Rate",
    "Required Flowrate": "Flow Rate", 
    "Set Pressure": "Set Pressure",
    "Pset": "Set Pressure", 
    "Built-up Back Pressure": "Built-up Back Pressure", 
    "Ratio of Max. Back Pressure": "Built-up Back Pressure", 
    "Relief Temperature": "Relief Temperature",
    "T": "Relief Temperature", 
    "Viscosity": "Viscosity",
    "mu": "Viscosity", 
    "Molecular Weight": "Molecular Weight",
    "M": "Molecular Weight", 
    "Gas Z": "Gas Z",
    "Z": "Gas Z", 
    "Max. BP@Header": "Max. BP@Header", 
    "Min. BP@Header": "Min. BP@Header", 
    "Relief Condition": "Relief Condition",
    "Rev. No.": "Rev. No.", 
    "Remark": "Remark",
    "Dwg No.": "Dwg No.",
    "Dwg. No.": "Dwg No.", 
    "PSV Type": "PSV Type",
    "No. of PSV": "No. of PSV",
    "No. of Installed PSV:": "No. of PSV", 
    "Allowable Overpressure": "Allowable Overpressure",
    "AllowOverPres": "Allowable Overpressure", 
    "Relief Case": "Relief Case",
    "PSV Material(CS/CMS/SS/NCA/A20)": "PSV Material",
    "WithRupDisk": "Installed with Rupture Disk", 
    "Density": "Density",
    "k": "Cp/Cv", 
    "Cp/Cv": "Cp/Cv",
    "TAG NO.": "Tag No.", "PHASE": "Phase", "FLOW RATE": "Flow Rate",
    "NORMAL PRESSURE": "Normal Pressure", "MECHANICAL DESIGN PRESSURE": "Mechanical Design Pressure",
    "SET PRESSURE": "Set Pressure", "CONST./VARIABLE SUPERIMPOSED BACK PRESSURE": "Const./Variable Superimposed Back Pressure",
    "BUILT-UP BACK PRESSURE": "Built-up Back Pressure", "FLARE SYSTEM": "Flare System",
    "ACCUMULATION": "Accumulation", "NORMAL TEMPERATURE": "Normal Temperature",
    "MECHANICAL DESIGN TEMPERATURE": "Mechanical Design Temperature", "RELIEF TEMPERATURE": "Relief Temperature",
    "VISCOSITY": "Viscosity", "MOLECULAR WEIGHT": "Molecular Weight", "GAS Z": "Gas Z",
    "MAX BP": "Max BP", "MIN BP": "Min BP", 
    "RELIEF CONDITION": "Relief Condition", 
    "REV. NO.": "Rev. No.", "REMARK": "Remark", "DWG NO.": "Dwg No.",
    "PSV TYPE": "PSV Type", "NO. OF PSV": "No. of PSV", "ALLOWABLE OVERPRESSURE": "Allowable Overpressure",
}

# --- Data Sheet Write Mapping ---
DATA_SHEET_WRITE_MAPPING = {
    (1, 0): ("Tag No.", None, None), 
    (4, 0): ("Relief Case", None, None), 
    (7, 0): ("Relief Condition", None, None), 
    (10, 0): ("Phase", get_state, None), 
    (11, 0): ("Flow Rate", None, None), 
    (12, 0): (None, None, 3.63), 
    (13, 0): (None, None, "50.8 / F.V."), 
    (14, 0): ("Set Pressure", None, None), 
    (15, 0): (("Min. BP@Header", "Max. BP@Header"), format_back_pressure_calculated, 'CALCULATED'), 
    (16, 0): (None, None, 4.12), 
    (17, 0): (None, None, "HVG"), 
    (18, 0): ("Allowable Overpressure", None, None), 
    (19, 0): (None, None, 104), 
    (20, 0): (None, None, 176), 
    (21, 0): ("Relief Temperature", None, None), 
    (22, 0): ("Viscosity", None, None), 
    (23, 0): ("Molecular Weight", None, None), 
    (24, 0): ("Gas Z", None, None), 
    (25, 0): (None, None, "0Ca"), 
    (26, 0): (None, None, "0Ca"), 
    (27, 0): ("Remark", None, None), 
    (1, 1): ("Dwg No.", None, None), 
    (10, 1): ("Fluid", None, None), 
    (17, 1): ("PSV Type", None, None), 
    (18, 1): (None, None, None), 
    (27, 1): ("Remark", None, None), 
}

# --- Fixed Unit Cells in Row 10 (to prevent clearing) ---
FIXED_UNIT_CELLS_IN_ROW_10 = [
    (11, 1), (12, 1), (13, 1), (14, 1), (15, 1), (16, 1),
    (19, 1), (20, 1), (21, 1), (22, 1), (23,1), (24,1)
]

def convert_calc_to_data_sheet(calc_sheet_raw_df, data_sheet_template_path, output_filename="Data_Sheet_filled_final.xlsm", output_stream=None):
    try:
        wb = openpyxl.load_workbook(data_sheet_template_path, keep_vba=True)
        ws = wb[DATA_SHEET_SHEET_NAME]
    except Exception as e:
        print(f"ERROR: Could not load Data Sheet template or sheet '{DATA_SHEET_SHEET_NAME}': {e}")
        return False

    try:
        psv_tag_nos_raw = calc_sheet_raw_df.iloc[1, 3:].tolist()
        psv_tag_nos = [
            str(x).strip() for x in psv_tag_nos_raw 
            if pd.notna(x) and str(x).strip() != ''
        ]
        
        if not psv_tag_nos:
            print("ERROR: No valid PSV Tag No. found in Calculation Sheet (expected in row 2, col D onwards).")
            return False

        prop_row_to_standard_name = {}
        for r_idx in range(1, calc_sheet_raw_df.shape[0]):
            raw_prop_name = calc_sheet_raw_df.iloc[r_idx, 1] 
            cleaned_name = str(raw_prop_name).strip() if pd.notna(raw_prop_name) else ''
            
            standardized_name = CALC_SHEET_HEADER_MAPPING.get(cleaned_name, None)
            
            if standardized_name is None:
                for key, value in CALC_SHEET_HEADER_MAPPING.items():
                    if key.strip().lower() == cleaned_name.lower():
                        standardized_name = value
                        break

            if standardized_name and standardized_name != "Tag No.": 
                prop_row_to_standard_name[r_idx] = standardized_name
        
        psv_records_list = []
        for col_offset, tag_no in enumerate(psv_tag_nos):
            current_psv_data = {'Tag No.': tag_no} 
            for raw_row_idx_in_df, standardized_name in prop_row_to_standard_name.items():
                if (3 + col_offset) < calc_sheet_raw_df.shape[1]:
                    value = calc_sheet_raw_df.iloc[raw_row_idx_in_df, 3 + col_offset]
                    current_psv_data[standardized_name] = value
            
            psv_records_list.append(current_psv_data)
        
        psv_records_df = pd.DataFrame(psv_records_list)

    except Exception as e:
        print(f"ERROR: Problem parsing Calculation Sheet data: {e}")
        return False
    
    all_mapped_excel_cols = [coord[0] for coord in DATA_SHEET_WRITE_MAPPING.keys()]

    min_col_to_clear = min(all_mapped_excel_cols) if all_mapped_excel_cols else 1
    max_col_to_clear = max(all_mapped_excel_cols) if all_mapped_excel_cols else 27

    original_merged_ranges = list(ws.merged_cells.ranges)
    for m_range_obj in original_merged_ranges:
        try:
            ws.unmerge_cells(str(m_range_obj))
        except Exception:
            pass

    for r in range(DATA_SHEET_FORM_START_ROW, MAX_ROWS_TO_CLEAR_IN_TEMPLATE + 1):
        for c in range(min_col_to_clear, max_col_to_clear + 1):
            cell = ws.cell(row=r, column=c)
            row_offset_in_pair = (r - DATA_SHEET_FORM_START_ROW) % 2
            
            is_fixed_unit_cell = False
            if (c, row_offset_in_pair) in FIXED_UNIT_CELLS_IN_ROW_10:
                is_fixed_unit_cell = True
                
            if not is_fixed_unit_cell:
                cell.value = None

    for idx, psv_record_row in psv_records_df.iterrows():
        current_excel_row_start = DATA_SHEET_FORM_START_ROW + idx * 2
        
        for (data_sheet_col, data_sheet_row_offset_in_pair), (calc_sheet_source, format_func, default_or_special) in DATA_SHEET_WRITE_MAPPING.items():
            
            value_to_write = None

            if default_or_special == 'CALCULATED':
                value_to_write = format_func(psv_record_row)
            elif default_or_special is not None:
                value_to_write = default_or_special
            elif calc_sheet_source is not None: 
                value_from_calc_sheet = psv_record_row.get(calc_sheet_source) 
                if format_func:
                    value_to_write = format_func(value_from_calc_sheet)
                else:
                    value_to_write = convert_value(value_from_calc_sheet)

            target_excel_row = current_excel_row_start + data_sheet_row_offset_in_pair
            cell = ws.cell(row=target_excel_row, column=data_sheet_col)
            cell.value = value_to_write

    for m_range_obj in original_merged_ranges:
        try:
            ws.merge_cells(str(m_range_obj))
        except Exception:
            pass

    try:
        if output_stream is not None:
            wb.save(output_stream)
            output_stream.seek(0)
            return output_stream
        else:
            wb.save(output_filename)
            return True
    except Exception as e:
        print(f"ERROR: Failed to save file '{output_filename}': {e}")
        return False

# --- Main Execution Block ---
if __name__ == '__main__':
    try:
        calc_sheet_raw_df = pd.read_excel("Calculation Sheet.xlsm", sheet_name=CALC_SHEET_SHEET_NAME, header=None)
        
        success = convert_calc_to_data_sheet(
            calc_sheet_raw_df, 
            DATA_SHEET_TEMPLATE_PATH,
            output_filename="Data_Sheet_filled_final.xlsm"
        )
        if not success:
            pass 

    except FileNotFoundError:
        print(f"ERROR: Calculation Sheet file '{'Calculation Sheet.xlsm'}' not found. Ensure the file exists in the same directory as the script.")
    except Exception as e:
        print(f"ERROR: An unexpected error occurred: {e}")